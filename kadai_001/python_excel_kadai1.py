# -*- coding: utf-8 -*-
"""python-excel-kadai1.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1qvMb1K8D3mwSsO7NxBP6CvcZfC-Lenom
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ワークブックの作成
wb = Workbook()
ws = wb.active

# データの挿入
ws['B2'] = '請求書'
ws.merge_cells('B2:E2')
ws['B4'] = '株式会社ABC'
ws.merge_cells('B4:E4')
ws['B5'] = '〒101-0022 東京都千代田区神田練塀町300'
ws.merge_cells('B5:E5')
ws['B6'] = 'TEL:03-1234-5678 FAX:03-1234-5678'
ws.merge_cells('B6:E6')
ws['B7'] = '担当者名:鈴木一郎 様'
ws.merge_cells('B7:E7')

# 日付の出力
ws['F4']= 'No.'
ws['G4']= '0001'
ws['F5']= '日付'
ws['G5'] = '=TEXT(TODAY(),"YYYY/MM/DD")'

# ヘッダーとデータの挿入
header = ['商品名', '数量', '単価', '金額']
data = [
    ['商品A', 2, 10000, 20000],
    ['商品B', 1, 15000, 15000]
]

ws['B10'] = header[0]
ws['C10'] = header[1]
ws['D10'] = header[2]
ws['E10'] = header[3]

for i, row in enumerate(data, start=11):
    for j, value in enumerate(row, start=2):
        cell = ws.cell(row=i, column=j, value=value)
        cell.alignment = Alignment(horizontal='center')

# 合計、消費税、税込み合計の挿入
ws['B15'] = '合計'
ws['B16'] = '消費税'
ws['B17'] = '税込み合計'

ws['E13'] = f'=SUM(E11:E12)'
ws['E15'] = f'=SUM(E11:E12)'
ws['E16'] = f'=E15*0.1'
ws['E17'] = f'=E15+E16'

# 列の幅を変更
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 10

# スタイルの設定
for row in ws.iter_rows(min_row=10, max_row=17):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

# ファイルの保存
wb.save('請求書.xlsx')