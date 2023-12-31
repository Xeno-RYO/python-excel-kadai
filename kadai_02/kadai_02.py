# -*- coding: utf-8 -*-
"""kadai_02.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1GL1wn1snQ0BBo9OvmvBO9xX_141O8yvo
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

header = ['日付', '社員名', '売上', '部門', '平均売上', '業績ランク']

# データフレームを作成する
df = pd.DataFrame(data={
    '日付': ['2023-05-17', '2023-05-18', '2023-05-19', '2023-05-20', '2023-05-21'],
    '社員名': ['山田', '佐藤', '鈴木', '田中', '高橋'],
    '売上': [100, 200, 150, 300, 250],
    '部門': ['メーカー', '代理店', 'メーカー', '商社', '代理店']
})

# 平均売上を計算し、列を追加する
average_sales = df['売上'].mean()
df['平均売上'] = average_sales

# 業績ランクを計算し、列を追加する
df['業績ランク'] = df['売上'].apply(lambda x: 'A' if x > average_sales else 'C' if x < average_sales else 'B')

# 結果を新しいExcelファイルとして保存する
df.to_excel('kadai_002.xlsx', index=False)

# Excelファイルを読み込む
book = load_workbook('kadai_002.xlsx')

# シートを選択する
worksheet = book['Sheet1']

# 列の幅を調整する
column_widths = [12, 12, 12, 12, 12, 12]
for col_num, width in enumerate(column_widths, 1):
    column_letter = chr(ord('A') + col_num - 1)
    worksheet.column_dimensions[column_letter].width = width

# ヘッダーのフォントを太字にする
for col_num, col_title in enumerate(header, 1):
    cell = worksheet.cell(row=1, column=col_num)
    cell.font = Font(bold=True)

# 枠線を追加する
border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.border = border

# Excelファイルを保存する
book.save('kadai_002.xlsx')